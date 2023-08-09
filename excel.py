from openpyxl import Workbook, load_workbook

# wb = Workbook()
# # Get active worksheet
# ws = wb.active
# print(ws)
# # Check default worksheet title
# ws.title = 'Default'

# # Create new sheets
# ws1 = wb.create_sheet('Test1')
# ws2 = wb.create_sheet('Test2')

# # Loop through all worksheets
# for sheet in wb:
#     print(sheet.title)

# # Using row and column notation
# d = ws.cell(row=4, column=2, value=10)

# # Assign a cell a value
# d.value = 3.14
# print(d.value)

# # Save Workbook file
# wb.save('test.xlsx')

# Load existing Excel file
wb = load_workbook(filename='test.xlsx')

ws['A1'].number_format = 'yyyy-mm-dd h:mm:ss'
